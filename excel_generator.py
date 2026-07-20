import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from datetime import datetime

def generate_excel(metadata, transactions, filepath, currency_symbol="₹"):
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement Analysis"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Style definitions
    font_family = "Segoe UI"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=11, bold=True, color="1B365D")
    meta_label_font = Font(name=font_family, size=10, bold=True, color="555555")
    meta_val_font = Font(name=font_family, size=10, color="000000")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, color="333333")
    total_font = Font(name=font_family, size=11, bold=True, color="000000")
    
    # KPI Fonts
    kpi_title_font = Font(name=font_family, size=9, bold=True, color="666666")
    kpi_val_font = Font(name=font_family, size=14, bold=True, color="1B365D")
    kpi_val_green_font = Font(name=font_family, size=14, bold=True, color="1E4620")
    kpi_val_red_font = Font(name=font_family, size=14, bold=True, color="7F1D1D")
    
    # Fills
    title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    kpi_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="CBD5E0")
    double_border_side = Side(border_style="double", color="1B365D")
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_bottom_double = Border(bottom=double_border_side, top=thin_border_side)
    border_top_thick = Border(top=thick_bottom_side)
    
    # 1. Title Block (Rows 1-2)
    ws.merge_cells("A1:G2")
    title_cell = ws["A1"]
    title_cell.value = "BANK STATEMENT ANALYSIS"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 2. Metadata Block (Rows 4-7)
    metadata_items = [
        ("Account Holder:", metadata.get("holder_name", "N/A")),
        ("Account Number:", metadata.get("account_number", "N/A")),
        ("Customer ID:", metadata.get("customer_id", "N/A")),
        ("Account Type:", metadata.get("account_type", "N/A")),
        ("Statement Period:", metadata.get("statement_period", "N/A")),
        ("Analysis Date:", datetime.now().strftime("%d-%b-%Y %I:%M %p"))
    ]
    
    row_idx = 4
    for label, val in metadata_items:
        ws.cell(row=row_idx, column=1, value=label).font = meta_label_font
        ws.cell(row=row_idx, column=2, value=val).font = meta_val_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left")
        row_idx += 1
        
    # Calculate stats
    total_debit = 0.0
    total_credit = 0.0
    
    for tx in transactions:
        try:
            d_val = float(str(tx.get("debit", "0")).replace(",", "").strip() or 0.0)
            total_debit += d_val
        except ValueError:
            pass
        try:
            c_val = float(str(tx.get("credit", "0")).replace(",", "").strip() or 0.0)
            total_credit += c_val
        except ValueError:
            pass
            
    net_flow = total_credit - total_debit
    
    # Get initial and final balance
    start_bal = 0.0
    end_bal = 0.0
    if transactions:
        try:
            start_bal = float(str(transactions[0].get("balance", "0")).replace(",", "").strip() or 0.0)
            # If the first row is B/F (brought forward) with balance but no transaction
            # we can use it.
        except ValueError:
            pass
        try:
            end_bal = float(str(transactions[-1].get("balance", "0")).replace(",", "").strip() or 0.0)
        except ValueError:
            pass

    # 3. KPI Blocks (Rows 4-7, columns D to G)
    # We will build 4 KPI Cards:
    # Card 1: Starting Balance (D4:D5)
    # Card 2: Ending Balance (E4:E5)
    # Card 3: Total Credit / Inflow (F4:F5)
    # Card 4: Total Debit / Outflow (G4:G5)
    
    def style_kpi_card(ws, start_col, title, value, is_amount=True, font_color_type="normal"):
        col_letter_1 = get_column_letter(start_col)
        col_letter_2 = get_column_letter(start_col)
        
        # Merge cell for KPI title
        ws.merge_cells(f"{col_letter_1}4:{col_letter_2}4")
        t_cell = ws[f"{col_letter_1}4"]
        t_cell.value = title.upper()
        t_cell.font = kpi_title_font
        t_cell.fill = kpi_fill
        t_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Merge cell for KPI value
        ws.merge_cells(f"{col_letter_1}5:{col_letter_2}6")
        v_cell = ws[f"{col_letter_1}5"]
        
        if is_amount:
            v_cell.value = value
            # Format pattern
            v_cell.number_format = f'"{currency_symbol}"#,##0.00'
        else:
            v_cell.value = value
            
        if font_color_type == "green":
            v_cell.font = kpi_val_green_font
        elif font_color_type == "red":
            v_cell.font = kpi_val_red_font
        else:
            v_cell.font = kpi_val_font
            
        v_cell.fill = kpi_fill
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Draw border around the card
        for r in range(4, 7):
            for c in range(start_col, start_col + 1):
                ws.cell(row=r, column=c).border = border_all
                
    style_kpi_card(ws, 4, "Starting Balance", start_bal, font_color_type="normal")
    style_kpi_card(ws, 5, "Ending Balance", end_bal, font_color_type="normal")
    style_kpi_card(ws, 6, "Total Credits (Inflow)", total_credit, font_color_type="green")
    style_kpi_card(ws, 7, "Total Debits (Outflow)", total_debit, font_color_type="red")
    
    # 4. Net Cash Flow Bar (Row 8, Column D to F merged)
    ws.merge_cells("D8:F8")
    net_cell = ws["D8"]
    net_cell.value = f"Net Cash Flow: "
    net_cell.font = Font(name=font_family, size=10, bold=True, color="333333")
    net_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    net_val_cell = ws.cell(row=8, column=7, value=net_flow)
    net_val_cell.font = Font(name=font_family, size=10, bold=True, color="1E4620" if net_flow >= 0 else "7F1D1D")
    net_val_cell.number_format = f'"{currency_symbol}"#,##0.00'
    net_val_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    for c in range(4, 8):
        ws.cell(row=8, column=c).border = border_all
        ws.cell(row=8, column=c).fill = kpi_fill
        
    # 5. Table Headers (Row 11)
    headers = ["Txn Date", "Value Date", "Description / Particulars", "Reference No.", "Debit", "Credit", "Balance"]
    header_row = 11
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_all
        
    ws.row_dimensions[header_row].height = 25
    
    # 6. Table Rows
    curr_row = header_row + 1
    
    number_format_str = f'#,##0.00'
    
    for tx in transactions:
        row_fill = zebra_fill if (curr_row % 2 == 0) else white_fill
        
        # Date columns
        c_date = ws.cell(row=curr_row, column=1, value=tx.get("txn_date", ""))
        c_val_date = ws.cell(row=curr_row, column=2, value=tx.get("value_date", ""))
        
        # Particulars
        c_part = ws.cell(row=curr_row, column=3, value=tx.get("particulars", ""))
        
        # Ref No
        c_ref = ws.cell(row=curr_row, column=4, value=tx.get("ref_no", ""))
        
        # Debit, Credit, Balance values (floats)
        try:
            d_val = float(str(tx.get("debit", "")).replace(",", "").strip() or 0.0)
            c_deb = ws.cell(row=curr_row, column=5, value=d_val if d_val > 0 else "")
            if d_val > 0:
                c_deb.number_format = number_format_str
        except ValueError:
            c_deb = ws.cell(row=curr_row, column=5, value=tx.get("debit", ""))
            
        try:
            c_val = float(str(tx.get("credit", "")).replace(",", "").strip() or 0.0)
            c_crd = ws.cell(row=curr_row, column=6, value=c_val if c_val > 0 else "")
            if c_val > 0:
                c_crd.number_format = number_format_str
        except ValueError:
            c_crd = ws.cell(row=curr_row, column=6, value=tx.get("credit", ""))
            
        try:
            b_val = float(str(tx.get("balance", "")).replace(",", "").strip() or 0.0)
            c_bal = ws.cell(row=curr_row, column=7, value=b_val)
            c_bal.number_format = number_format_str
        except ValueError:
            c_bal = ws.cell(row=curr_row, column=7, value=tx.get("balance", ""))
            
        # Alignments
        c_date.alignment = Alignment(horizontal="center", vertical="center")
        c_val_date.alignment = Alignment(horizontal="center", vertical="center")
        c_part.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c_ref.alignment = Alignment(horizontal="center", vertical="center")
        c_deb.alignment = Alignment(horizontal="right", vertical="center")
        c_crd.alignment = Alignment(horizontal="right", vertical="center")
        c_bal.alignment = Alignment(horizontal="right", vertical="center")
        
        # Apply fonts, borders, and fills
        for col_idx in range(1, 8):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = border_all
            
        curr_row += 1
        
    # 7. Total Row
    total_row = curr_row
    
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
    tot_label = ws.cell(row=total_row, column=1, value="Total")
    tot_label.font = total_font
    tot_label.alignment = Alignment(horizontal="right", vertical="center")
    
    tot_deb = ws.cell(row=total_row, column=5, value=total_debit)
    tot_deb.font = total_font
    tot_deb.number_format = number_format_str
    tot_deb.alignment = Alignment(horizontal="right", vertical="center")
    
    tot_crd = ws.cell(row=total_row, column=6, value=total_credit)
    tot_crd.font = total_font
    tot_crd.number_format = number_format_str
    tot_crd.alignment = Alignment(horizontal="right", vertical="center")
    
    # Empty balance cell for totals (or write end balance)
    tot_bal = ws.cell(row=total_row, column=7, value="")
    
    for col_idx in range(1, 8):
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = total_fill
        cell.border = border_bottom_double
        
    ws.row_dimensions[total_row].height = 20
    
    # 8. Column Width Adjustments
    # Auto-adjust column widths based on content lengths
    min_widths = [15, 15, 45, 18, 15, 15, 18] # Desired min width for each column
    
    for col_idx in range(1, 8):
        col_letter = get_column_letter(col_idx)
        # Find maximum length of cell value in the table data
        max_len = min_widths[col_idx - 1]
        
        # Check table cells (from row 11 to total_row)
        for r in range(11, total_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val:
                # If description is very long, wrap it and limit width to 45
                if col_idx == 3:
                    max_len = 45
                else:
                    max_len = max(max_len, len(str(val)) + 3)
                    
        ws.column_dimensions[col_letter].width = max_len
        
    # Set layout adjustments for metadata columns
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 30
    
    # Save workbook
    wb.save(filepath)
